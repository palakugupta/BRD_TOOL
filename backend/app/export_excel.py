"""
BRD Quality Assessment – Excel Report Generator (stable + formatted)
"""

import sqlite3
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─────────────────────────────────────────────
# Colors
# ─────────────────────────────────────────────

BRAND_DARK = "1E2A3A"
BRAND_LIGHT = "E8F0FE"
WHITE = "FFFFFF"

RED_BG = "FDE8E8"
YELLOW_BG = "FFF8E1"
GREEN_BG = "E8F5E9"

RED_ACCENT = "C62828"
YELLOW_ACCENT = "F57F17"
GREEN_ACCENT = "2E7D32"


SEVERITY_COLOURS = {
    "critical": (RED_BG, RED_ACCENT),
    "major": (YELLOW_BG, YELLOW_ACCENT),
    "minor": (GREEN_BG, GREEN_ACCENT),
}


ERROR_TYPES = [
    "different_data",
    "incomplete_data",
    "hallucination",
    "depth_differs",
    "duplicate_data",
]


ERROR_LABELS = {
    "different_data": "Different Data",
    "incomplete_data": "Incomplete Data",
    "hallucination": "Hallucination",
    "depth_differs": "Depth Mismatch",
    "duplicate_data": "Duplicate Data",
}


# ─────────────────────────────────────────────
# Styles
# ─────────────────────────────────────────────

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)


CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=16, color=WHITE)


# ─────────────────────────────────────────────
# Fetch data
# ─────────────────────────────────────────────

def _fetch_findings(conn):

    sql = """
        SELECT
            f.finding_id,
            f.chunk_id,
            f.error_type,
            f.severity,
            f.line_number,
            f.description,
            f.source_reference,
            f.rule_id,
            f.detected_timestamp,
            c.start_line,
            c.end_line
        FROM findings f
        LEFT JOIN chunks c
        ON f.chunk_id = c.chunk_id
        ORDER BY f.line_number
    """

    cur = conn.execute(sql)

    rows = cur.fetchall()

    cols = [c[0] for c in cur.description]

    return [dict(zip(cols, r)) for r in rows]


def _fetch_coverage(conn):

    total = conn.execute(
        "SELECT COALESCE(SUM(line_count),0) FROM documents WHERE doc_type='output_brd'"
    ).fetchone()[0]

    flagged = conn.execute(
        """
        SELECT COUNT(DISTINCT line_number)
        FROM findings
        WHERE line_number IS NOT NULL
        AND line_number > 0
        """
    ).fetchone()[0]

    total = total or 0
    flagged = flagged or 0

    coverage = ((total - flagged) / total * 100) if total else 0

    return total, flagged, coverage


def _fetch_run_meta(conn):

    def latest(doc_type):

        row = conn.execute(
            "SELECT filename FROM documents WHERE doc_type=? ORDER BY doc_id DESC LIMIT 1",
            (doc_type,),
        ).fetchone()

        return row[0] if row else "N/A"

    return {
        "sow": latest("input_sow"),
        "mom": latest("input_mom"),
        "brd": latest("output_brd"),
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ─────────────────────────────────────────────
# Summary sheet
# ─────────────────────────────────────────────

def _build_summary(ws, findings, meta, total_lines, flagged_lines, coverage):

    ws.title = "Summary"

    ws.merge_cells("A1:F1")

    title = ws["A1"]
    title.value = "BRD Quality Assessment Report"
    title.font = TITLE_FONT
    title.fill = _fill(BRAND_DARK)
    title.alignment = CENTER

    meta_rows = [
        ("SOW File", meta["sow"]),
        ("MoM File", meta["mom"]),
        ("BRD File", meta["brd"]),
        ("Generated On", meta["generated"]),
        ("Total BRD Lines", total_lines),
        ("Lines Flagged", flagged_lines),
    ]

    row = 3

    for label, val in meta_rows:
        ws.cell(row, 1, label).font = Font(bold=True)
        ws.cell(row, 2, val)
        row += 1

    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    cov = ws.cell(row, 1)

    cov.value = f"Coverage Score: {coverage:.1f}%"

    if coverage >= 90:
        cov.fill = _fill(GREEN_BG)
    elif coverage >= 70:
        cov.fill = _fill(YELLOW_BG)
    else:
        cov.fill = _fill(RED_BG)

    cov.alignment = CENTER
    cov.font = Font(bold=True)

    row += 2

    headers = ["Error Type", "Total", "Critical", "Major", "Minor"]

    for i, h in enumerate(headers, 1):

        cell = ws.cell(row, i, h)

        cell.font = HEADER_FONT
        cell.fill = _fill(BRAND_DARK)
        cell.alignment = CENTER
        cell.border = _border()

    summary = {}

    for f in findings:

        et = f["error_type"]
        sv = f["severity"]

        summary.setdefault(et, {"critical": 0, "major": 0, "minor": 0})

        if sv in summary[et]:
            summary[et][sv] += 1

    row += 1

    for etype in ERROR_TYPES:

        counts = summary.get(etype, {"critical": 0, "major": 0, "minor": 0})

        total = sum(counts.values())

        ws.cell(row, 1, ERROR_LABELS[etype])
        ws.cell(row, 2, total)
        ws.cell(row, 3, counts["critical"])
        ws.cell(row, 4, counts["major"])
        ws.cell(row, 5, counts["minor"])

        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12


# ─────────────────────────────────────────────
# Detail sheets
# ─────────────────────────────────────────────

DETAIL_HEADERS = [
    "Finding ID",
    "Line #",
    "Chunk Range",
    "Severity",
    "Description",
    "Source Reference",
]


def _build_detail_sheet(ws, label, findings):

    ws.title = label[:31]

    ws.append(DETAIL_HEADERS)

    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = _fill(BRAND_DARK)
        c.alignment = CENTER
        c.border = _border()

    for f in findings:

        start = f.get("start_line")
        end = f.get("end_line")

        chunk_range = f"{start}-{end}" if start and end else ""

        row = [
            f.get("finding_id"),
            f.get("line_number"),
            chunk_range,
            f.get("severity"),
            f.get("description"),
            f.get("source_reference"),
        ]

        ws.append(row)

        r = ws.max_row

        sev = (f.get("severity") or "").lower()

        bg = SEVERITY_COLOURS.get(sev, (WHITE,))[0]

        for c in ws[r]:

            c.fill = _fill(bg)
            c.border = _border()
            c.alignment = LEFT
            c.font = BODY_FONT

    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 30


# ─────────────────────────────────────────────
# All findings sheet
# ─────────────────────────────────────────────

ALL_HEADERS = [
    "Finding ID",
    "Error Type",
    "Line #",
    "Chunk Range",
    "Severity",
    "Description",
]


def _build_all_sheet(ws, findings):

    ws.title = "All Findings"

    ws.append(ALL_HEADERS)

    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = _fill(BRAND_DARK)
        c.alignment = CENTER
        c.border = _border()

    for f in findings:

        start = f.get("start_line")
        end = f.get("end_line")

        chunk_range = f"{start}-{end}" if start and end else ""

        row = [
            f.get("finding_id"),
            ERROR_LABELS.get(f.get("error_type"), f.get("error_type")),
            f.get("line_number"),
            chunk_range,
            f.get("severity"),
            f.get("description"),
        ]

        ws.append(row)

        r = ws.max_row

        sev = (f.get("severity") or "").lower()

        bg = SEVERITY_COLOURS.get(sev, (WHITE,))[0]

        for c in ws[r]:
            c.fill = _fill(bg)
            c.border = _border()
            c.alignment = LEFT

    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 60


# ─────────────────────────────────────────────
# Main generator
# ─────────────────────────────────────────────

def generate_excel_report(db_path: str) -> bytes:

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    findings = _fetch_findings(conn)

    total, flagged, coverage = _fetch_coverage(conn)

    meta = _fetch_run_meta(conn)

    wb = Workbook()

    ws = wb.active

    _build_summary(ws, findings, meta, total, flagged, coverage)

    for etype in ERROR_TYPES:

        subset = [f for f in findings if f["error_type"] == etype]

        ws = wb.create_sheet(ERROR_LABELS[etype])

        _build_detail_sheet(ws, ERROR_LABELS[etype], subset)

    ws = wb.create_sheet()

    _build_all_sheet(ws, findings)

    conn.close()

    buffer = io.BytesIO()

    wb.save(buffer)

    return buffer.getvalue()